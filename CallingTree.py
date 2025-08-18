# --- Required Libraries ---
import streamlit as st
import pandas as pd
from io import BytesIO
import datetime
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
import os

# --- Streamlit Page Configuration ---
st.set_page_config(page_title="Calling Tree Login", layout="wide")

# --- Login Credentials (Hardcoded) ---
USERNAME = "Backoffice"
PASSWORD = "3NOv-51bT-Q8iu"


def extract_date(file_name):
    # Extract the "Mon-YY" part and parse it
    date_str = file_name.split(" - ")[1]
    return datetime.datetime.strptime(date_str, "%b-%y")


# --- Session State Initialization ---
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

# --- Login Page Function ---
def login_page():
    st.title("🔐 Calling Tree Login")
    username = st.text_input("Username").replace(" ", "")
    password = st.text_input("Password", type="password").replace(" ", "")
    
    if st.button("Login"):
        if username == USERNAME and password == PASSWORD:
            st.session_state.logged_in = True
            st.rerun()
        # elif username == "custom" or password == "":
        #     st.session_state.logged_in = True
        #     st.rerun()
        else:
            st.error("❌ Incorrect username or password.")




# --- Function to Load Data from SharePoint ---
def get_Calling_Tree(file_name="Calling Tree - Aug-25.xlsx"):
    try:
        # Simulate loading data from SharePoint
        # Replace this with actual SharePoint file loading logic
        df = pd.read_excel("Calling_Trees/"+file_name)
        return df
    except Exception as e:
        st.error(f"Error loading data: {e}")
        return pd.DataFrame()  # Return an empty DataFrame on error



# --- Main App Logic ---
if st.session_state.logged_in:


    # get all xlsx files from folder 
    folder_path = "Calling_Trees"
    xlsx_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
    
    if not xlsx_files:     
        st.error("No Calling Tree files found in the folder.")
        st.stop()
    # Display the calling tree selection interface
    xlsx_files = [os.path.splitext(f)[0] for f in xlsx_files]

    # Sort by date descending
    xlsx_files = sorted(xlsx_files, key=extract_date, reverse=True)

    col1, col2 = st.columns([9,1])
    with col1:
        st.title("📞 Calling Tree Management")  

    with col2:
        st.markdown("")
        st.markdown("")
        if st.button("🔓 Logout"):
            st.session_state.logged_in = False
            st.rerun()

    col2, col3 = st.columns([1, 1])

    with col2:
        option = st.selectbox("Select Month", xlsx_files)
        calling_tree_df = get_Calling_Tree(option+".xlsx")



    with col3:
        st.markdown('<div style="font-size:17px;">&nbsp;</div>', unsafe_allow_html=True)
    

        file_month = datetime.datetime.now().strftime("%b-%y")
        filename = option+".xlsx"

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            calling_tree_df.to_excel(writer, sheet_name="Main", index=False)
            workbook = writer.book
            worksheet = writer.sheets["Main"]

            header_fill = PatternFill(start_color="0C769E", end_color="0C769E", fill_type="solid")
            body_fill = PatternFill(start_color="CAEDFB", end_color="CAEDFB", fill_type="solid")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            align_center = Alignment(horizontal="center", vertical="center")
            bold_white = Font(bold=True, color="FFFFFF")

            # Header styling
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = bold_white
                cell.alignment = align_center
                cell.border = border

            # Data styling
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
                for cell in row:
                    cell.fill = body_fill
                    cell.alignment = align_center
                    cell.border = border

        output.seek(0)
        st.download_button(
            label="⬇️ Download Calling Tree",
            data=output,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )




    # Display the editable DataFrame
    st.data_editor(
        calling_tree_df,
        use_container_width=True,
        hide_index=True,
        height=720
    )
else:
    login_page()
